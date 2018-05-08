# -*- coding: utf-8 -*-

import os
import json
import openpyxl

def search(dirname):
    """이 함수는 디렉토리(폴더) 경로를 받아 해당 디렉토리의 파일 리스트를 반환함.

    : param: str dirname: 디렉토리(폴더) 경로

    예제:
        다음과 같이 사용:
        >>> filelist = search("C:/temp")
            print(filelist)
            ['C:/temp/test1', 'C:/temp/test2', 'C:/temp/test3']
 
    :returns list: 파일경로 리스트 반환
    """

    fullFilenames = []
    filenames = os.listdir(dirname)
    for filename in filenames:
        fullFilename = os.path.join(dirname, filename)
        fullFilenames.append(fullFilename)

    # print (fullFilenames)

    return fullFilenames

def jsonRead(filename):
    """이 함수는 json 파일을 읽어 python dictionary 값으로 반환.

    : param: str filename: 파일 경로.

    예제:
        다음과 같이 사용:
        >>> secuCheckList = jsonRead("SecuCheckList.json")

    :returns dict: json 파일 내용을 dictionary 타입 값으로 반환
    """
    fileData = open(filename, "r").read()
    jsonDataAsPythonValue = json.loads(fileData)
    return jsonDataAsPythonValue

# 결과 파일 경로 리스트
fullFilenames = search("201805")

#print(fullFilenames)
#print(type(fullFilenames))

# 점검 항목 리스트
secuCheckDict = jsonRead("SecuCheckList.json")
secuCheckList = list(secuCheckDict.items())

#print(secuCheckDict)
#print(type(secuCheckDict))
#print(len(secuCheckList))

# 엑셀 파일 초기화
wb = openpyxl.Workbook()
ws = wb.active
ws.cell(row=1, column=1, value=u"코드")
ws.cell(row=1, column=2, value=u"항목명")
ws.cell(row=1, column=3, value=u"양호")
ws.cell(row=1, column=4, value=u"취약")
ws.cell(row=1, column=5, value=u"예외")

for row in range(2, len(secuCheckList)+2):
    #코드, 항목명 입력
    for col in range(1, 3):
        ws.cell(row=row, column=col, value=secuCheckList[row-2][col-1])
    #항목별 양호, 취약, 예외 개수 확인 함수 입력
    lineTrueCount = "=countif(" + "F" + str(row) + ":DF" + str(row) + ", C1)"
    lineFalseCount = "=countif(" + "F" + str(row) + ":DF" + str(row) + ", D1)"
    lineNullCount = "=countif(" + "F" + str(row) + ":DF" + str(row) + ", E1)"
    ws.cell(row=row, column=3, value=lineTrueCount)
    ws.cell(row=row, column=4, value=lineFalseCount)
    ws.cell(row=row, column=5, value=lineNullCount)
wb.save("/Users/luminary_topco/Desktop/201805-Topco_PC_Security_Check.xlsx")
wb.close()

# 엑셀 파일에 점검 결과 입력
wb = openpyxl.load_workbook("./201805-Topco_PC_Security_Check.xlsx")
ws = wb.active

colNum = 6
for fullFilename in fullFilenames:
    checkCount =[0, 0, 0]
    secuCheckData = jsonRead(fullFilename)
    ws.cell(row=1, column=colNum, value=secuCheckData['name'])
    for row in range(2, len(secuCheckList)+2):
        val = secuCheckData[secuCheckList[row-2][0]]
        if val == "ture": # 진단 스크립트 오타로 결과 수정
            val = "true"
        if val == "true":
            checkCount[0] += 1
            val = u"양호"
        elif val == "false":
            checkCount[1] += 1
            val = u"취약"
        elif val == "null":
            checkCount[2] += 1
            val = u"예외"

        ws.cell(row=row, column=colNum, value=val)

        #print(secuCheckData[secuCheckList[row-2][0]])

    # 결과 개 수
    checkCountStr = u"양호=" + str(checkCount[0]) + u"/ 취약=" + str(checkCount[1]) + u"/ 에외=" + str(checkCount[2])
    ws.cell(row=len(secuCheckList)+2, column=colNum, value=checkCountStr)

    # true 백분률
    r = ((checkCount[0] + checkCount[2]) / len(secuCheckList)) * 100
    ws.cell(row=len(secuCheckList)+3, column=colNum, value=r)

    colNum += 1

    #print(colNum)

wb.save("./201805-Topco_PC_Security_Check.xlsx")
wb.close()

#    print(secuCheckData)
#    print(type(secuCheckData))
